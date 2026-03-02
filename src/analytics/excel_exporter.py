# src/analytics/excel_exporter.py

"""
Excel Data Verification Exporter

Generates Excel workbooks for data science team verification of generated content.
Includes:
- Sheet 1: Figure verification (extracted metrics from markdown)
- Sheet 2: Query log (all database queries executed)
- Sheet 3: Calculation trace (derived metrics with formulas)
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import re
import json
from datetime import datetime
from typing import Dict, List, Any, Optional, Callable
from dataclasses import dataclass, field
from functools import wraps
import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from config.settings import settings


@dataclass
class QueryLogEntry:
    """A logged database query"""
    query: str
    params: Dict[str, Any]
    execution_time_ms: float
    row_count: int
    timestamp: datetime
    source_function: str
    result_sample: Optional[List[Dict]] = None


@dataclass
class CalculationEntry:
    """A logged calculation/derivation"""
    metric_name: str
    formula: str
    input_values: Dict[str, Any]
    output_value: Any
    source_function: str
    timestamp: datetime
    notes: Optional[str] = None


@dataclass
class FigureEntry:
    """A figure extracted from generated content"""
    value: Any
    formatted_value: str
    context: str  # Surrounding text
    metric_type: str  # 'count', 'price', 'percentage', etc.
    source_query: Optional[str] = None
    validation_status: str = "pending"  # 'pass', 'fail', 'pending'
    validation_note: Optional[str] = None


class QueryLogger:
    """
    Context manager and decorator for logging database queries.

    Usage:
        # As context manager
        with QueryLogger() as logger:
            result = db.execute("SELECT * FROM table")
            logger.log_query("SELECT * FROM table", {}, result)

        # As decorator
        @QueryLogger.track_queries
        def my_function(db):
            return db.execute("SELECT * FROM table")
    """

    _instance = None
    _entries: List[QueryLogEntry] = []
    _enabled: bool = False

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    @classmethod
    def enable(cls):
        """Enable query logging"""
        cls._enabled = True
        cls._entries = []

    @classmethod
    def disable(cls):
        """Disable query logging"""
        cls._enabled = False

    @classmethod
    def clear(cls):
        """Clear logged entries"""
        cls._entries = []

    @classmethod
    def get_entries(cls) -> List[QueryLogEntry]:
        """Get all logged entries"""
        return cls._entries.copy()

    @classmethod
    def log_query(
        cls,
        query: str,
        params: Dict[str, Any],
        execution_time_ms: float,
        row_count: int,
        source_function: str,
        result_sample: Optional[List[Dict]] = None
    ):
        """Log a database query"""
        if not cls._enabled:
            return

        entry = QueryLogEntry(
            query=query.strip(),
            params=params or {},
            execution_time_ms=execution_time_ms,
            row_count=row_count,
            timestamp=datetime.now(),
            source_function=source_function,
            result_sample=result_sample
        )
        cls._entries.append(entry)

    @classmethod
    def track_queries(cls, func: Callable) -> Callable:
        """Decorator to track queries from a function"""
        @wraps(func)
        def wrapper(*args, **kwargs):
            start_time = time.time()
            result = func(*args, **kwargs)
            execution_time = (time.time() - start_time) * 1000

            # Try to extract query info if result has it
            if hasattr(result, 'description'):
                row_count = result.rowcount if hasattr(result, 'rowcount') else -1
                cls.log_query(
                    query="[Auto-tracked query]",
                    params={},
                    execution_time_ms=execution_time,
                    row_count=row_count,
                    source_function=func.__name__
                )

            return result
        return wrapper


class CalculationLogger:
    """
    Logger for tracking calculations and derived metrics.

    Usage:
        logger = CalculationLogger()
        logger.log_calculation(
            metric_name="avg_price",
            formula="SUM(price) / COUNT(*)",
            input_values={"sum": 1000000, "count": 10},
            output_value=100000,
            source_function="calculate_avg_price"
        )
    """

    _instance = None
    _entries: List[CalculationEntry] = []
    _enabled: bool = False

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    @classmethod
    def enable(cls):
        """Enable calculation logging"""
        cls._enabled = True
        cls._entries = []

    @classmethod
    def disable(cls):
        """Disable calculation logging"""
        cls._enabled = False

    @classmethod
    def clear(cls):
        """Clear logged entries"""
        cls._entries = []

    @classmethod
    def get_entries(cls) -> List[CalculationEntry]:
        """Get all logged entries"""
        return cls._entries.copy()

    @classmethod
    def log_calculation(
        cls,
        metric_name: str,
        formula: str,
        input_values: Dict[str, Any],
        output_value: Any,
        source_function: str,
        notes: Optional[str] = None
    ):
        """Log a calculation"""
        if not cls._enabled:
            return

        entry = CalculationEntry(
            metric_name=metric_name,
            formula=formula,
            input_values=input_values,
            output_value=output_value,
            source_function=source_function,
            timestamp=datetime.now(),
            notes=notes
        )
        cls._entries.append(entry)


class ExcelExporter:
    """
    Generates Excel verification workbooks for data science team.

    Usage:
        exporter = ExcelExporter()

        # Add figures extracted from content
        exporter.add_figure(value=1234, formatted="1,234", context="Total transactions: 1,234")

        # Add query logs
        exporter.add_query_log(QueryLogger.get_entries())

        # Add calculation traces
        exporter.add_calculation_log(CalculationLogger.get_entries())

        # Export to Excel
        exporter.export("verification_report.xlsx")
    """

    # Style definitions
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    PENDING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, report_name: str = None):
        """
        Initialize the exporter.

        Args:
            report_name: Name of the report being verified
        """
        self.report_name = report_name or "Verification Report"
        self.figures: List[FigureEntry] = []
        self.query_logs: List[QueryLogEntry] = []
        self.calculation_logs: List[CalculationEntry] = []
        self.metadata: Dict[str, Any] = {
            "generated_at": datetime.now().isoformat(),
            "report_name": self.report_name
        }

    def extract_figures_from_markdown(self, markdown_content: str) -> List[FigureEntry]:
        """
        Extract all figures/numbers from markdown content.

        Identifies:
        - Transaction counts (1,234 transactions)
        - Prices (AED 1,234,567)
        - Percentages (12.5%)
        - Areas (1,234 sqm)
        - Year references (2024, 2025)

        Args:
            markdown_content: The generated markdown content

        Returns:
            List of FigureEntry objects
        """
        figures = []

        # Pattern definitions with context extraction
        patterns = [
            # AED prices (with commas)
            (r'AED\s*([\d,]+(?:\.\d+)?)', 'price', 'AED {value}'),
            # Percentage values
            (r'([\d.]+)\s*%', 'percentage', '{value}%'),
            # Transaction/unit counts with comma formatting
            (r'([\d,]+)\s+(?:transactions?|units?|properties|sales)', 'count', '{value} units'),
            # Square meters
            (r'([\d,]+(?:\.\d+)?)\s*(?:sqm|sq\.?\s*m)', 'area', '{value} sqm'),
            # Standalone large numbers (likely counts/volumes)
            (r'(?<![.\d])([\d]{1,3}(?:,\d{3})+)(?![.\d%])', 'large_number', '{value}'),
        ]

        # Track already found values to avoid duplicates
        found_values = set()

        for pattern, metric_type, format_template in patterns:
            for match in re.finditer(pattern, markdown_content, re.IGNORECASE):
                raw_value = match.group(1)

                # Parse numeric value
                try:
                    numeric_value = float(raw_value.replace(',', ''))
                except ValueError:
                    continue

                # Skip if already found
                value_key = (numeric_value, metric_type)
                if value_key in found_values:
                    continue
                found_values.add(value_key)

                # Extract surrounding context (50 chars each side)
                start = max(0, match.start() - 50)
                end = min(len(markdown_content), match.end() + 50)
                context = markdown_content[start:end].replace('\n', ' ').strip()

                # Format the value
                formatted = format_template.format(value=raw_value)

                figures.append(FigureEntry(
                    value=numeric_value,
                    formatted_value=formatted,
                    context=f"...{context}...",
                    metric_type=metric_type
                ))

        self.figures.extend(figures)
        return figures

    def add_figure(
        self,
        value: Any,
        formatted_value: str,
        context: str,
        metric_type: str = "unknown",
        source_query: Optional[str] = None,
        validation_status: str = "pending",
        validation_note: Optional[str] = None
    ):
        """Add a figure entry manually"""
        self.figures.append(FigureEntry(
            value=value,
            formatted_value=formatted_value,
            context=context,
            metric_type=metric_type,
            source_query=source_query,
            validation_status=validation_status,
            validation_note=validation_note
        ))

    def add_query_log(self, entries: List[QueryLogEntry]):
        """Add query log entries"""
        self.query_logs.extend(entries)

    def add_calculation_log(self, entries: List[CalculationEntry]):
        """Add calculation log entries"""
        self.calculation_logs.extend(entries)

    def set_metadata(self, key: str, value: Any):
        """Set metadata for the report"""
        self.metadata[key] = value

    def _apply_header_style(self, ws, row: int, num_cols: int):
        """Apply header styling to a row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_data_style(self, ws, start_row: int, end_row: int, num_cols: int):
        """Apply data styling to rows"""
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

    def _auto_column_width(self, ws, num_cols: int, max_width: int = 50):
        """Auto-adjust column widths"""
        for col in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col)

            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = min(cell_length, max_width)
                except:
                    pass

            ws.column_dimensions[column_letter].width = max(max_length + 2, 10)

    def _create_figures_sheet(self, wb: Workbook):
        """Create the Figure Verification sheet"""
        ws = wb.create_sheet("Figure Verification")

        # Headers
        headers = [
            "ID", "Value", "Formatted", "Type", "Context",
            "Source Query", "Status", "Validation Note"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))

        # Data
        for idx, fig in enumerate(self.figures, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=fig.value)
            ws.cell(row=row, column=3, value=fig.formatted_value)
            ws.cell(row=row, column=4, value=fig.metric_type)
            ws.cell(row=row, column=5, value=fig.context)
            ws.cell(row=row, column=6, value=fig.source_query or "")
            ws.cell(row=row, column=7, value=fig.validation_status)
            ws.cell(row=row, column=8, value=fig.validation_note or "")

            # Apply status coloring
            status_cell = ws.cell(row=row, column=7)
            if fig.validation_status == "pass":
                status_cell.fill = self.PASS_FILL
            elif fig.validation_status == "fail":
                status_cell.fill = self.FAIL_FILL
            else:
                status_cell.fill = self.PENDING_FILL

        self._apply_data_style(ws, 2, len(self.figures) + 1, len(headers))
        self._auto_column_width(ws, len(headers))

    def _create_query_log_sheet(self, wb: Workbook):
        """Create the Query Log sheet"""
        ws = wb.create_sheet("Query Log")

        # Headers
        headers = [
            "ID", "Timestamp", "Source Function", "Query",
            "Parameters", "Execution Time (ms)", "Row Count"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))

        # Data
        for idx, entry in enumerate(self.query_logs, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=3, value=entry.source_function)
            ws.cell(row=row, column=4, value=entry.query[:500] if len(entry.query) > 500 else entry.query)
            ws.cell(row=row, column=5, value=json.dumps(entry.params) if entry.params else "")
            ws.cell(row=row, column=6, value=round(entry.execution_time_ms, 2))
            ws.cell(row=row, column=7, value=entry.row_count)

        self._apply_data_style(ws, 2, len(self.query_logs) + 1, len(headers))
        self._auto_column_width(ws, len(headers))

    def _create_calculation_sheet(self, wb: Workbook):
        """Create the Calculation Trace sheet"""
        ws = wb.create_sheet("Calculation Trace")

        # Headers
        headers = [
            "ID", "Timestamp", "Metric Name", "Formula",
            "Input Values", "Output Value", "Source Function", "Notes"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))

        # Data
        for idx, entry in enumerate(self.calculation_logs, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=3, value=entry.metric_name)
            ws.cell(row=row, column=4, value=entry.formula)
            ws.cell(row=row, column=5, value=json.dumps(entry.input_values, default=str))
            ws.cell(row=row, column=6, value=str(entry.output_value))
            ws.cell(row=row, column=7, value=entry.source_function)
            ws.cell(row=row, column=8, value=entry.notes or "")

        self._apply_data_style(ws, 2, len(self.calculation_logs) + 1, len(headers))
        self._auto_column_width(ws, len(headers))

    def _create_summary_sheet(self, wb: Workbook):
        """Create a summary/metadata sheet"""
        ws = wb.active
        ws.title = "Summary"

        # Title
        ws.cell(row=1, column=1, value="Data Verification Report")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')

        # Metadata
        ws.cell(row=3, column=1, value="Report Name:")
        ws.cell(row=3, column=2, value=self.report_name)

        ws.cell(row=4, column=1, value="Generated At:")
        ws.cell(row=4, column=2, value=self.metadata.get("generated_at", ""))

        ws.cell(row=5, column=1, value="Total Figures:")
        ws.cell(row=5, column=2, value=len(self.figures))

        ws.cell(row=6, column=1, value="Total Queries:")
        ws.cell(row=6, column=2, value=len(self.query_logs))

        ws.cell(row=7, column=1, value="Total Calculations:")
        ws.cell(row=7, column=2, value=len(self.calculation_logs))

        # Validation summary
        ws.cell(row=9, column=1, value="Validation Summary")
        ws.cell(row=9, column=1).font = Font(bold=True, size=12)

        pass_count = sum(1 for f in self.figures if f.validation_status == "pass")
        fail_count = sum(1 for f in self.figures if f.validation_status == "fail")
        pending_count = sum(1 for f in self.figures if f.validation_status == "pending")

        ws.cell(row=10, column=1, value="Passed:")
        ws.cell(row=10, column=2, value=pass_count)
        ws.cell(row=10, column=2).fill = self.PASS_FILL

        ws.cell(row=11, column=1, value="Failed:")
        ws.cell(row=11, column=2, value=fail_count)
        ws.cell(row=11, column=2).fill = self.FAIL_FILL

        ws.cell(row=12, column=1, value="Pending:")
        ws.cell(row=12, column=2, value=pending_count)
        ws.cell(row=12, column=2).fill = self.PENDING_FILL

        # Additional metadata
        row = 14
        ws.cell(row=row, column=1, value="Additional Metadata")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)

        for key, value in self.metadata.items():
            if key not in ["generated_at", "report_name"]:
                row += 1
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))

        # Auto-width
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20

    def export(self, filepath: str = None) -> Path:
        """
        Export verification data to Excel workbook.

        Args:
            filepath: Output path. If None, generates default path.

        Returns:
            Path to the generated Excel file
        """
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = re.sub(r'[^\w\-]', '_', self.report_name)
            filepath = settings.CONTENT_OUTPUT_DIR / f"verification_{safe_name}_{timestamp}.xlsx"
        else:
            filepath = Path(filepath)

        # Ensure directory exists
        filepath.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Create sheets
        self._create_summary_sheet(wb)
        self._create_figures_sheet(wb)
        self._create_query_log_sheet(wb)
        self._create_calculation_sheet(wb)

        # Save
        wb.save(str(filepath))

        return filepath


def create_verification_report(
    markdown_content: str,
    report_name: str,
    output_path: str = None,
    validation_results: Dict[str, Any] = None
) -> Path:
    """
    Convenience function to create a verification report from generated content.

    Args:
        markdown_content: The generated markdown content
        report_name: Name of the report
        output_path: Optional output path
        validation_results: Optional validation results to include

    Returns:
        Path to the generated Excel file
    """
    exporter = ExcelExporter(report_name=report_name)

    # Extract figures
    exporter.extract_figures_from_markdown(markdown_content)

    # Add query logs if available
    exporter.add_query_log(QueryLogger.get_entries())

    # Add calculation logs if available
    exporter.add_calculation_log(CalculationLogger.get_entries())

    # Add validation results if provided
    if validation_results:
        exporter.set_metadata("validation_pass_rate", validation_results.get("pass_rate", "N/A"))
        exporter.set_metadata("validation_tests_passed", validation_results.get("passed", 0))
        exporter.set_metadata("validation_tests_total", validation_results.get("total_tests", 0))

    return exporter.export(output_path)


if __name__ == "__main__":
    # Test the exporter
    print("Testing ExcelExporter...")

    # Sample markdown content
    test_markdown = """
    # Dubai Market Report Q4 2024

    ## Overview
    Total transactions: 15,234 in Q4 2024, up 12.5% from Q3.

    ## Price Analysis
    Average price: AED 2,345,678 per unit.
    Price per sqm: AED 12,500 per sqm.

    ## Market Segments
    - Luxury segment (5M+): 2,345 transactions (15.4%)
    - Off-plan sales: 8,901 units (58.4%)
    """

    exporter = ExcelExporter(report_name="Test Report Q4 2024")

    # Extract figures
    figures = exporter.extract_figures_from_markdown(test_markdown)
    print(f"Extracted {len(figures)} figures")
    for fig in figures:
        print(f"  - {fig.formatted_value} ({fig.metric_type})")

    # Add some test query logs
    QueryLogger.enable()
    QueryLogger.log_query(
        query="SELECT COUNT(*) FROM transactions WHERE year = 2024",
        params={"year": 2024},
        execution_time_ms=45.2,
        row_count=1,
        source_function="get_transaction_count"
    )
    exporter.add_query_log(QueryLogger.get_entries())

    # Add some test calculations
    CalculationLogger.enable()
    CalculationLogger.log_calculation(
        metric_name="avg_price",
        formula="SUM(price) / COUNT(*)",
        input_values={"sum": 23456780000, "count": 10000},
        output_value=2345678,
        source_function="calculate_avg_price"
    )
    exporter.add_calculation_log(CalculationLogger.get_entries())

    # Export
    output_path = exporter.export()
    print(f"\nGenerated: {output_path}")
