# scripts/generate_2025_annual_report.py

"""
2025 Dubai Real Estate Annual Market Report Generator

Generates:
1. Full market report content with AI
2. Comprehensive verification document for data science team
3. Excel workbook with all figures, SQL queries, and calculation methods

Usage:
    python scripts/generate_2025_annual_report.py
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import duckdb
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict

from config.bigquery_settings import bq_settings
from config.settings import settings

from src.analytics.report_calculator import ReportCalculator, PeriodType
from src.analytics.market_intelligence import (
    MarketIntelligenceEngine,
    AnomalyDetector,
    OpportunityDetector,
    TrendPredictor,
    ComparativeAnalytics
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


@dataclass
class SQLQuery:
    """A documented SQL query for verification"""
    name: str
    description: str
    sql: str
    parameters: Dict[str, Any]
    result_count: int
    result_sample: Optional[List[Dict]] = None
    notes: Optional[str] = None


@dataclass
class CalculatedMetric:
    """A calculated metric with full derivation"""
    name: str
    value: Any
    formatted_value: str
    formula: str
    inputs: Dict[str, Any]
    source_module: str
    category: str  # 'transaction', 'price', 'area', 'developer', 'intelligence'


class Report2025Generator:
    """
    Generates 2025 Annual Market Report with full verification documentation.
    """

    def __init__(self):
        self.db_path = bq_settings.PM_DB_PATH
        self.con = duckdb.connect(str(self.db_path), read_only=True)
        self.output_dir = settings.CONTENT_OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Trackers for verification
        self.sql_queries: List[SQLQuery] = []
        self.metrics: List[CalculatedMetric] = []

        # Report parameters
        self.year = 2025
        self.period_type = 'annual'
        self.period_number = 1

    def _log_query(self, name: str, description: str, sql: str, params: Dict = None,
                   result_count: int = 0, result_sample: List = None, notes: str = None):
        """Log a SQL query for verification"""
        self.sql_queries.append(SQLQuery(
            name=name,
            description=description,
            sql=sql.strip(),
            parameters=params or {},
            result_count=result_count,
            result_sample=result_sample,
            notes=notes
        ))

    def _log_metric(self, name: str, value: Any, formatted: str, formula: str,
                    inputs: Dict, source: str, category: str):
        """Log a calculated metric for verification"""
        self.metrics.append(CalculatedMetric(
            name=name,
            value=value,
            formatted_value=formatted,
            formula=formula,
            inputs=inputs,
            source_module=source,
            category=category
        ))

    def generate_core_metrics(self) -> Dict[str, Any]:
        """Generate and log all core metrics with their SQL queries."""
        print("\n" + "="*60)
        print("GENERATING CORE METRICS")
        print("="*60)

        results = {}

        # =====================================================================
        # 1. TOTAL TRANSACTIONS AND VALUE
        # =====================================================================
        sql = """
            SELECT
                COUNT(*) as total_transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price,
                MEDIAN(actual_worth) as median_price,
                MIN(actual_worth) as min_price,
                MAX(actual_worth) as max_price
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
        """
        row = self.con.execute(sql).fetchone()

        self._log_query(
            name="total_transactions_2025",
            description="Total transaction count and value metrics for 2025",
            sql=sql,
            params={"start_date": "2025-01-01", "end_date": "2026-01-01"},
            result_count=1,
            notes="Core aggregate metrics for the annual report"
        )

        results['total_transactions'] = row[0]
        results['total_value'] = row[1]
        results['avg_price'] = row[2]
        results['median_price'] = row[3]
        results['min_price'] = row[4]
        results['max_price'] = row[5]

        self._log_metric("Total Transactions", row[0], f"{row[0]:,}", "COUNT(*)",
                        {}, "report_calculator", "transaction")
        self._log_metric("Total Value", row[1], f"AED {row[1]:,.0f}", "SUM(actual_worth)",
                        {}, "report_calculator", "transaction")
        self._log_metric("Average Price", row[2], f"AED {row[2]:,.0f}", "AVG(actual_worth)",
                        {}, "report_calculator", "price")
        self._log_metric("Median Price", row[3], f"AED {row[3]:,.0f}", "MEDIAN(actual_worth)",
                        {}, "report_calculator", "price")

        print(f"  Total Transactions: {row[0]:,}")
        print(f"  Total Value: AED {row[1]:,.0f}")
        print(f"  Average Price: AED {row[2]:,.0f}")

        # =====================================================================
        # 2. YEAR-OVER-YEAR COMPARISON
        # =====================================================================
        sql_yoy = """
            SELECT
                EXTRACT(YEAR FROM instance_date::DATE) as year,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price
            FROM transactions_clean
            WHERE instance_date >= '2024-01-01' AND instance_date < '2026-01-01'
            GROUP BY EXTRACT(YEAR FROM instance_date::DATE)
            ORDER BY year
        """
        yoy_df = self.con.execute(sql_yoy).df()

        self._log_query(
            name="yoy_comparison",
            description="Year-over-year comparison 2024 vs 2025",
            sql=sql_yoy,
            result_count=len(yoy_df),
            result_sample=yoy_df.to_dict('records'),
            notes="Used to calculate YoY growth rates"
        )

        if len(yoy_df) >= 2:
            tx_2024 = yoy_df[yoy_df['year'] == 2024]['transactions'].values[0]
            tx_2025 = yoy_df[yoy_df['year'] == 2025]['transactions'].values[0]
            val_2024 = yoy_df[yoy_df['year'] == 2024]['total_value'].values[0]
            val_2025 = yoy_df[yoy_df['year'] == 2025]['total_value'].values[0]
            avg_2024 = yoy_df[yoy_df['year'] == 2024]['avg_price'].values[0]
            avg_2025 = yoy_df[yoy_df['year'] == 2025]['avg_price'].values[0]

            tx_growth = ((tx_2025 - tx_2024) / tx_2024) * 100
            val_growth = ((val_2025 - val_2024) / val_2024) * 100
            price_growth = ((avg_2025 - avg_2024) / avg_2024) * 100

            results['yoy'] = {
                'tx_2024': tx_2024, 'tx_2025': tx_2025, 'tx_growth': tx_growth,
                'val_2024': val_2024, 'val_2025': val_2025, 'val_growth': val_growth,
                'avg_2024': avg_2024, 'avg_2025': avg_2025, 'price_growth': price_growth
            }

            self._log_metric("Transaction Growth YoY", tx_growth, f"{tx_growth:+.1f}%",
                           f"((tx_2025 - tx_2024) / tx_2024) * 100 = (({tx_2025} - {tx_2024}) / {tx_2024}) * 100",
                           {"tx_2025": tx_2025, "tx_2024": tx_2024}, "report_calculator", "transaction")
            self._log_metric("Value Growth YoY", val_growth, f"{val_growth:+.1f}%",
                           f"((val_2025 - val_2024) / val_2024) * 100",
                           {"val_2025": val_2025, "val_2024": val_2024}, "report_calculator", "transaction")
            self._log_metric("Price Growth YoY", price_growth, f"{price_growth:+.1f}%",
                           f"((avg_2025 - avg_2024) / avg_2024) * 100",
                           {"avg_2025": avg_2025, "avg_2024": avg_2024}, "report_calculator", "price")

            print(f"\n  YoY Transaction Growth: {tx_growth:+.1f}%")
            print(f"  YoY Value Growth: {val_growth:+.1f}%")
            print(f"  YoY Price Growth: {price_growth:+.1f}%")

        # =====================================================================
        # 3. MONTHLY BREAKDOWN
        # =====================================================================
        sql_monthly = """
            SELECT
                EXTRACT(MONTH FROM instance_date::DATE) as month,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
            GROUP BY EXTRACT(MONTH FROM instance_date::DATE)
            ORDER BY month
        """
        monthly_df = self.con.execute(sql_monthly).df()

        self._log_query(
            name="monthly_breakdown_2025",
            description="Monthly transaction breakdown for 2025",
            sql=sql_monthly,
            result_count=len(monthly_df),
            result_sample=monthly_df.to_dict('records'),
            notes="Shows monthly distribution throughout 2025"
        )

        results['monthly'] = monthly_df.to_dict('records')
        print(f"\n  Monthly data: {len(monthly_df)} months captured")

        # =====================================================================
        # 4. PROPERTY TYPE BREAKDOWN
        # =====================================================================
        sql_prop_type = """
            SELECT
                property_type_en,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) as percentage
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
              AND property_type_en IS NOT NULL
            GROUP BY property_type_en
            ORDER BY transactions DESC
        """
        prop_df = self.con.execute(sql_prop_type).df()

        self._log_query(
            name="property_type_breakdown",
            description="Transaction breakdown by property type",
            sql=sql_prop_type,
            result_count=len(prop_df),
            result_sample=prop_df.to_dict('records')
        )

        results['property_types'] = prop_df.to_dict('records')
        for _, row in prop_df.iterrows():
            self._log_metric(f"Property Type: {row['property_type_en']}",
                           row['transactions'],
                           f"{row['transactions']:,} ({row['percentage']:.1f}%)",
                           "COUNT(*) WHERE property_type_en = X",
                           {"property_type": row['property_type_en']},
                           "report_calculator", "transaction")

        print(f"\n  Property Types: {len(prop_df)} categories")

        # =====================================================================
        # 5. OFF-PLAN VS READY (REG_TYPE_EN)
        # =====================================================================
        sql_offplan = """
            SELECT
                reg_type_en,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) as percentage
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
              AND reg_type_en IS NOT NULL
            GROUP BY reg_type_en
            ORDER BY transactions DESC
        """
        offplan_df = self.con.execute(sql_offplan).df()

        self._log_query(
            name="offplan_vs_ready",
            description="Off-plan vs Ready property breakdown",
            sql=sql_offplan,
            result_count=len(offplan_df),
            result_sample=offplan_df.to_dict('records'),
            notes="reg_type_en distinguishes off-plan from ready properties"
        )

        results['market_segments'] = offplan_df.to_dict('records')
        print(f"\n  Market Segments (Off-plan/Ready): {len(offplan_df)} categories")

        # =====================================================================
        # 6. TOP 10 AREAS BY TRANSACTION COUNT
        # =====================================================================
        sql_areas = """
            SELECT
                area_name_en,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price,
                AVG(meter_sale_price) as avg_price_sqm,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM transactions_clean
                    WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'), 2) as market_share
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
              AND area_name_en IS NOT NULL
            GROUP BY area_name_en
            ORDER BY transactions DESC
            LIMIT 20
        """
        areas_df = self.con.execute(sql_areas).df()

        self._log_query(
            name="top_areas_2025",
            description="Top 20 areas by transaction count in 2025",
            sql=sql_areas,
            result_count=len(areas_df),
            result_sample=areas_df.to_dict('records')
        )

        results['top_areas'] = areas_df.to_dict('records')
        for i, row in areas_df.head(10).iterrows():
            self._log_metric(f"Area #{i+1}: {row['area_name_en']}",
                           row['transactions'],
                           f"{row['transactions']:,} ({row['market_share']:.1f}%)",
                           "COUNT(*) WHERE area_name_en = X",
                           {"area": row['area_name_en'], "market_share": row['market_share']},
                           "report_calculator", "area")

        print(f"\n  Top Areas: {len(areas_df)} areas analyzed")

        # =====================================================================
        # 7. TOP 10 DEVELOPERS
        # =====================================================================
        sql_developers = """
            SELECT
                master_project_en as developer,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM transactions_clean
                    WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
                    AND master_project_en IS NOT NULL), 2) as market_share
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
              AND master_project_en IS NOT NULL
              AND master_project_en != ''
            GROUP BY master_project_en
            ORDER BY transactions DESC
            LIMIT 20
        """
        dev_df = self.con.execute(sql_developers).df()

        self._log_query(
            name="top_developers_2025",
            description="Top 20 developers by transaction count in 2025",
            sql=sql_developers,
            result_count=len(dev_df),
            result_sample=dev_df.to_dict('records')
        )

        results['top_developers'] = dev_df.to_dict('records')
        print(f"\n  Top Developers: {len(dev_df)} developers analyzed")

        # =====================================================================
        # 8. PRICE SEGMENTS
        # =====================================================================
        sql_segments = """
            SELECT
                CASE
                    WHEN actual_worth < 1000000 THEN 'Under 1M'
                    WHEN actual_worth < 2000000 THEN '1M-2M'
                    WHEN actual_worth < 5000000 THEN '2M-5M'
                    WHEN actual_worth < 10000000 THEN '5M-10M'
                    WHEN actual_worth < 20000000 THEN '10M-20M'
                    ELSE '20M+'
                END as price_segment,
                COUNT(*) as transactions,
                SUM(actual_worth) as total_value,
                AVG(actual_worth) as avg_price,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 2) as percentage
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
            GROUP BY
                CASE
                    WHEN actual_worth < 1000000 THEN 'Under 1M'
                    WHEN actual_worth < 2000000 THEN '1M-2M'
                    WHEN actual_worth < 5000000 THEN '2M-5M'
                    WHEN actual_worth < 10000000 THEN '5M-10M'
                    WHEN actual_worth < 20000000 THEN '10M-20M'
                    ELSE '20M+'
                END
            ORDER BY
                CASE price_segment
                    WHEN 'Under 1M' THEN 1
                    WHEN '1M-2M' THEN 2
                    WHEN '2M-5M' THEN 3
                    WHEN '5M-10M' THEN 4
                    WHEN '10M-20M' THEN 5
                    ELSE 6
                END
        """
        segments_df = self.con.execute(sql_segments).df()

        self._log_query(
            name="price_segments_2025",
            description="Transaction distribution by price segment",
            sql=sql_segments,
            result_count=len(segments_df),
            result_sample=segments_df.to_dict('records'),
            notes="Price segments: Under 1M, 1M-2M, 2M-5M, 5M-10M (Luxury), 10M-20M, 20M+ (Ultra-luxury)"
        )

        results['price_segments'] = segments_df.to_dict('records')
        print(f"\n  Price Segments: {len(segments_df)} segments")

        # =====================================================================
        # 9. RECORD TRANSACTIONS
        # =====================================================================
        sql_records = """
            SELECT
                transaction_id,
                instance_date,
                area_name_en,
                property_type_en,
                rooms_en,
                actual_worth,
                meter_sale_price,
                master_project_en
            FROM transactions_clean
            WHERE instance_date >= '2025-01-01' AND instance_date < '2026-01-01'
            ORDER BY actual_worth DESC
            LIMIT 10
        """
        records_df = self.con.execute(sql_records).df()

        self._log_query(
            name="record_transactions_2025",
            description="Top 10 highest-value transactions in 2025",
            sql=sql_records,
            result_count=len(records_df),
            result_sample=records_df.to_dict('records'),
            notes="Record-breaking transactions for headline findings"
        )

        results['record_transactions'] = records_df.to_dict('records')
        print(f"\n  Record Transactions: Top {len(records_df)} captured")

        return results

    def generate_intelligence_metrics(self) -> Dict[str, Any]:
        """Generate market intelligence metrics with full documentation."""
        print("\n" + "="*60)
        print("GENERATING MARKET INTELLIGENCE")
        print("="*60)

        results = {}

        # Initialize intelligence modules
        anomaly = AnomalyDetector(self.con)
        opportunity = OpportunityDetector(self.con)
        trend = TrendPredictor(self.con)
        comparative = ComparativeAnalytics(self.con)

        # =====================================================================
        # 1. ANOMALY DETECTION
        # =====================================================================
        print("\n  Running Anomaly Detection...")
        anomalies = anomaly.detect_anomalies(self.year, self.period_type, self.period_number)

        results['anomalies'] = {
            'record_transactions': [
                {'area': tx.area, 'price': tx.price, 'type': tx.property_type, 'rank': tx.rank}
                for tx in anomalies.record_transactions
            ],
            'volume_spikes': [
                {'area': s.area, 'current_volume': s.current_volume,
                 'historical_avg': s.historical_avg, 'spike_pct': s.spike_percentage}
                for s in anomalies.volume_spikes
            ],
            'new_developers': [
                {'developer': d.developer, 'first_sale': d.first_sale_date,
                 'transactions': d.transaction_count, 'value': d.total_value}
                for d in anomalies.new_developers
            ]
        }

        # Log anomaly queries
        self._log_query(
            name="anomaly_volume_spikes",
            description="Areas with transaction volume significantly above historical average",
            sql="[See AnomalyDetector._get_volume_spikes method]",
            params={"lookback_months": 6, "spike_threshold": 30},
            result_count=len(anomalies.volume_spikes),
            notes="Compares current period volume to 6-month historical monthly average"
        )

        print(f"    Record transactions: {len(anomalies.record_transactions)}")
        print(f"    Volume spikes: {len(anomalies.volume_spikes)}")
        print(f"    New developers: {len(anomalies.new_developers)}")

        # =====================================================================
        # 2. OPPORTUNITY DETECTION
        # =====================================================================
        print("\n  Running Opportunity Detection...")
        opportunities = opportunity.detect_opportunities(self.year, self.period_type, self.period_number)

        results['opportunities'] = {
            'emerging_hotspots': [
                {'area': h.area, 'tx_growth': h.tx_growth_rate,
                 'price_growth': h.price_growth_rate, 'current_volume': h.current_tx_count}
                for h in opportunities.emerging_hotspots
            ],
            'undervalued_areas': [
                {'area': a.area, 'discount': a.price_discount,
                 'avg_price_sqm': a.avg_price_sqm, 'comparable_avg': a.comparable_avg_price_sqm}
                for a in opportunities.undervalued_areas
            ],
            'developer_momentum': [
                {'developer': d.developer, 'market_share_change': d.market_share_change,
                 'trend': d.trend, 'current_share': d.current_market_share}
                for d in opportunities.developer_momentum
            ]
        }

        print(f"    Emerging hotspots: {len(opportunities.emerging_hotspots)}")
        print(f"    Undervalued areas: {len(opportunities.undervalued_areas)}")
        print(f"    Developer momentum: {len(opportunities.developer_momentum)}")

        # =====================================================================
        # 3. TREND PREDICTION
        # =====================================================================
        print("\n  Running Trend Analysis...")
        trends = trend.analyze_trends()

        results['trends'] = {
            'price_momentum': {
                '3m': trends.price_momentum.momentum_3m,
                '6m': trends.price_momentum.momentum_6m,
                '12m': trends.price_momentum.momentum_12m,
                'trajectory': trends.price_momentum.trajectory.value
            },
            'volume_momentum': trends.volume_momentum,
            'seasonality': {
                'peak_months': trends.seasonality.peak_months,
                'low_months': trends.seasonality.low_months
            },
            'cycle_position': {
                'phase': trends.cycle_position.phase.value,
                'confidence': trends.cycle_position.phase_confidence
            }
        }

        if trends.price_momentum.momentum_12m:
            self._log_metric("12-Month Price Momentum", trends.price_momentum.momentum_12m,
                           f"{trends.price_momentum.momentum_12m:+.1f}%",
                           "((current_3m_avg - prior_12m_avg) / prior_12m_avg) * 100",
                           {}, "trend_predictor", "intelligence")

        print(f"    Price trajectory: {trends.price_momentum.trajectory.value}")
        print(f"    Market cycle: {trends.cycle_position.phase.value}")

        # =====================================================================
        # 4. COMPARATIVE ANALYTICS
        # =====================================================================
        print("\n  Running Comparative Analytics...")
        concentration = comparative.get_market_concentration()

        results['market_concentration'] = concentration

        if concentration:
            self._log_metric("Top 5 Areas Market Share",
                           concentration['area_concentration']['top_5_share'],
                           f"{concentration['area_concentration']['top_5_share']:.1f}%",
                           "SUM(top_5_area_transactions) / total_transactions * 100",
                           {}, "comparative_analytics", "intelligence")

        print(f"    Market concentration data generated")

        return results

    def generate_verification_excel(self, core_metrics: Dict, intelligence: Dict) -> Path:
        """Generate comprehensive Excel verification workbook."""
        print("\n" + "="*60)
        print("GENERATING VERIFICATION EXCEL")
        print("="*60)

        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # =====================================================================
        # SHEET 1: EXECUTIVE SUMMARY
        # =====================================================================
        ws1 = wb.active
        ws1.title = "Executive Summary"

        summary_data = [
            ["2025 Dubai Real Estate Market Report - Verification Document"],
            [""],
            ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Database:", str(self.db_path)],
            ["Report Period:", "Annual 2025 (Jan 1 - Dec 31)"],
            [""],
            ["KEY METRICS"],
            ["Metric", "Value", "YoY Change"],
            ["Total Transactions", f"{core_metrics['total_transactions']:,}",
             f"{core_metrics['yoy']['tx_growth']:+.1f}%" if 'yoy' in core_metrics else "N/A"],
            ["Total Value", f"AED {core_metrics['total_value']:,.0f}",
             f"{core_metrics['yoy']['val_growth']:+.1f}%" if 'yoy' in core_metrics else "N/A"],
            ["Average Price", f"AED {core_metrics['avg_price']:,.0f}",
             f"{core_metrics['yoy']['price_growth']:+.1f}%" if 'yoy' in core_metrics else "N/A"],
            ["Median Price", f"AED {core_metrics['median_price']:,.0f}", "N/A"],
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 8:
                    cell.font = header_font
                    cell.fill = header_fill

        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 25
        ws1.column_dimensions['C'].width = 15

        # =====================================================================
        # SHEET 2: ALL METRICS
        # =====================================================================
        ws2 = wb.create_sheet("All Metrics")

        headers = ["Metric Name", "Value", "Formatted", "Formula", "Source Module", "Category"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, metric in enumerate(self.metrics, 2):
            ws2.cell(row=row_idx, column=1, value=metric.name).border = border
            ws2.cell(row=row_idx, column=2, value=str(metric.value)).border = border
            ws2.cell(row=row_idx, column=3, value=metric.formatted_value).border = border
            ws2.cell(row=row_idx, column=4, value=metric.formula).border = border
            ws2.cell(row=row_idx, column=5, value=metric.source_module).border = border
            ws2.cell(row=row_idx, column=6, value=metric.category).border = border

        for col in range(1, 7):
            ws2.column_dimensions[get_column_letter(col)].width = 25

        print(f"  Metrics sheet: {len(self.metrics)} metrics logged")

        # =====================================================================
        # SHEET 3: SQL QUERIES
        # =====================================================================
        ws3 = wb.create_sheet("SQL Queries")

        headers = ["Query Name", "Description", "SQL Query", "Parameters", "Row Count", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, query in enumerate(self.sql_queries, 2):
            ws3.cell(row=row_idx, column=1, value=query.name).border = border
            ws3.cell(row=row_idx, column=2, value=query.description).border = border
            ws3.cell(row=row_idx, column=3, value=query.sql[:500] + "..." if len(query.sql) > 500 else query.sql).border = border
            ws3.cell(row=row_idx, column=4, value=json.dumps(query.parameters)).border = border
            ws3.cell(row=row_idx, column=5, value=query.result_count).border = border
            ws3.cell(row=row_idx, column=6, value=query.notes or "").border = border

        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 40
        ws3.column_dimensions['C'].width = 80
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 40

        print(f"  SQL Queries sheet: {len(self.sql_queries)} queries logged")

        # =====================================================================
        # SHEET 4: MONTHLY BREAKDOWN
        # =====================================================================
        ws4 = wb.create_sheet("Monthly Data")

        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']

        headers = ["Month", "Transactions", "Total Value (AED)", "Avg Price (AED)"]
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, month_data in enumerate(core_metrics['monthly'], 2):
            month_num = int(month_data['month'])
            ws4.cell(row=row_idx, column=1, value=month_names[month_num]).border = border
            ws4.cell(row=row_idx, column=2, value=month_data['transactions']).border = border
            ws4.cell(row=row_idx, column=3, value=month_data['total_value']).border = border
            ws4.cell(row=row_idx, column=4, value=month_data['avg_price']).border = border

        for col in range(1, 5):
            ws4.column_dimensions[get_column_letter(col)].width = 20

        # =====================================================================
        # SHEET 5: TOP AREAS
        # =====================================================================
        ws5 = wb.create_sheet("Top Areas")

        headers = ["Rank", "Area", "Transactions", "Total Value (AED)", "Avg Price (AED)", "Avg Price/sqm", "Market Share %"]
        for col, header in enumerate(headers, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, area in enumerate(core_metrics['top_areas'], 2):
            ws5.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws5.cell(row=row_idx, column=2, value=area['area_name_en']).border = border
            ws5.cell(row=row_idx, column=3, value=area['transactions']).border = border
            ws5.cell(row=row_idx, column=4, value=area['total_value']).border = border
            ws5.cell(row=row_idx, column=5, value=area['avg_price']).border = border
            ws5.cell(row=row_idx, column=6, value=area.get('avg_price_sqm', 0)).border = border
            ws5.cell(row=row_idx, column=7, value=area['market_share']).border = border

        for col in range(1, 8):
            ws5.column_dimensions[get_column_letter(col)].width = 18

        # =====================================================================
        # SHEET 6: TOP DEVELOPERS
        # =====================================================================
        ws6 = wb.create_sheet("Top Developers")

        headers = ["Rank", "Developer", "Transactions", "Total Value (AED)", "Avg Price (AED)", "Market Share %"]
        for col, header in enumerate(headers, 1):
            cell = ws6.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, dev in enumerate(core_metrics['top_developers'], 2):
            ws6.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws6.cell(row=row_idx, column=2, value=dev['developer']).border = border
            ws6.cell(row=row_idx, column=3, value=dev['transactions']).border = border
            ws6.cell(row=row_idx, column=4, value=dev['total_value']).border = border
            ws6.cell(row=row_idx, column=5, value=dev['avg_price']).border = border
            ws6.cell(row=row_idx, column=6, value=dev['market_share']).border = border

        for col in range(1, 7):
            ws6.column_dimensions[get_column_letter(col)].width = 20

        # =====================================================================
        # SHEET 7: PRICE SEGMENTS
        # =====================================================================
        ws7 = wb.create_sheet("Price Segments")

        headers = ["Price Segment", "Transactions", "Total Value (AED)", "Avg Price (AED)", "Percentage %"]
        for col, header in enumerate(headers, 1):
            cell = ws7.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, seg in enumerate(core_metrics['price_segments'], 2):
            ws7.cell(row=row_idx, column=1, value=seg['price_segment']).border = border
            ws7.cell(row=row_idx, column=2, value=seg['transactions']).border = border
            ws7.cell(row=row_idx, column=3, value=seg['total_value']).border = border
            ws7.cell(row=row_idx, column=4, value=seg['avg_price']).border = border
            ws7.cell(row=row_idx, column=5, value=seg['percentage']).border = border

        for col in range(1, 6):
            ws7.column_dimensions[get_column_letter(col)].width = 20

        # =====================================================================
        # SHEET 8: RECORD TRANSACTIONS
        # =====================================================================
        ws8 = wb.create_sheet("Record Transactions")

        headers = ["Rank", "Transaction ID", "Date", "Area", "Property Type", "Rooms", "Price (AED)", "Price/sqm", "Developer"]
        for col, header in enumerate(headers, 1):
            cell = ws8.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, tx in enumerate(core_metrics['record_transactions'], 2):
            ws8.cell(row=row_idx, column=1, value=row_idx - 1).border = border
            ws8.cell(row=row_idx, column=2, value=tx['transaction_id']).border = border
            ws8.cell(row=row_idx, column=3, value=str(tx['instance_date'])).border = border
            ws8.cell(row=row_idx, column=4, value=tx['area_name_en']).border = border
            ws8.cell(row=row_idx, column=5, value=tx['property_type_en']).border = border
            ws8.cell(row=row_idx, column=6, value=tx.get('rooms_en', 'N/A')).border = border
            ws8.cell(row=row_idx, column=7, value=tx['actual_worth']).border = border
            ws8.cell(row=row_idx, column=8, value=tx.get('meter_sale_price', 0)).border = border
            ws8.cell(row=row_idx, column=9, value=tx.get('master_project_en', 'N/A')).border = border

        for col in range(1, 10):
            ws8.column_dimensions[get_column_letter(col)].width = 18

        # =====================================================================
        # SHEET 9: MARKET INTELLIGENCE
        # =====================================================================
        ws9 = wb.create_sheet("Market Intelligence")

        intel_rows = [
            ["MARKET INTELLIGENCE ANALYSIS"],
            [""],
            ["ANOMALY DETECTION"],
            ["Volume Spikes Detected:", len(intelligence.get('anomalies', {}).get('volume_spikes', []))],
            ["New Developers Detected:", len(intelligence.get('anomalies', {}).get('new_developers', []))],
            [""],
            ["OPPORTUNITY DETECTION"],
            ["Emerging Hotspots:", len(intelligence.get('opportunities', {}).get('emerging_hotspots', []))],
            ["Undervalued Areas:", len(intelligence.get('opportunities', {}).get('undervalued_areas', []))],
            [""],
            ["TREND ANALYSIS"],
            ["Price Trajectory:", intelligence.get('trends', {}).get('price_momentum', {}).get('trajectory', 'N/A')],
            ["Market Cycle Phase:", intelligence.get('trends', {}).get('cycle_position', {}).get('phase', 'N/A')],
            [""],
            ["MARKET CONCENTRATION"],
        ]

        if intelligence.get('market_concentration'):
            conc = intelligence['market_concentration']
            intel_rows.append(["Top 5 Areas Share:", f"{conc['area_concentration']['top_5_share']:.1f}%"])
            intel_rows.append(["Top 5 Developers Share:", f"{conc['developer_concentration']['top_5_share']:.1f}%"])

        for row_idx, row_data in enumerate(intel_rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws9.cell(row=row_idx, column=col_idx, value=str(value))
                if row_idx in [1, 3, 7, 11, 15]:
                    cell.font = Font(bold=True)

        ws9.column_dimensions['A'].width = 25
        ws9.column_dimensions['B'].width = 30

        # Save workbook
        output_path = self.output_dir / f"2025_annual_report_verification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(output_path)

        print(f"\n  Verification Excel saved: {output_path}")
        return output_path

    def run(self):
        """Run the full report generation process."""
        print("\n" + "="*60)
        print("2025 DUBAI REAL ESTATE ANNUAL REPORT GENERATOR")
        print("="*60)
        print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Database: {self.db_path}")

        # Generate core metrics
        core_metrics = self.generate_core_metrics()

        # Generate intelligence metrics
        intelligence = self.generate_intelligence_metrics()

        # Generate verification Excel
        excel_path = self.generate_verification_excel(core_metrics, intelligence)

        # Save raw data as JSON for reference
        json_path = self.output_dir / f"2025_annual_report_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'core_metrics': core_metrics,
                'intelligence': intelligence,
                'sql_queries': [asdict(q) for q in self.sql_queries],
                'metrics': [asdict(m) for m in self.metrics]
            }, f, indent=2, default=str)

        print(f"\n  Raw data JSON saved: {json_path}")

        print("\n" + "="*60)
        print("GENERATION COMPLETE")
        print("="*60)
        print(f"\nOutputs:")
        print(f"  1. Verification Excel: {excel_path}")
        print(f"  2. Raw Data JSON: {json_path}")
        print(f"\nNext Steps:")
        print(f"  1. Review the verification Excel with the data science team")
        print(f"  2. Once verified, run the full report with AI content generation")

        return {
            'excel_path': excel_path,
            'json_path': json_path,
            'core_metrics': core_metrics,
            'intelligence': intelligence
        }


if __name__ == "__main__":
    generator = Report2025Generator()
    results = generator.run()
